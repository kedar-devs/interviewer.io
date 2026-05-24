import asyncio
import os

import certifi

os.environ["SSL_CERT_FILE"] = certifi.where()
os.environ["REQUESTS_CA_BUNDLE"] = certifi.where()

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from autogen_agentchat.agents import AssistantAgent, UserProxyAgent
from autogen_agentchat.conditions import (
    MaxMessageTermination,
    TextMentionTermination,
)
from autogen_agentchat.teams import SelectorGroupChat
from autogen_agentchat.ui import Console

from autogen_ext.models.openai import OpenAIChatCompletionClient
from autogen_ext.tools.mcp import McpWorkbench, StdioServerParams
from pathlib import Path

load_dotenv()

# if not os.getenv("OPENAI_API_KEY"):
#     raise RuntimeError(
#         "OPENAI_API_KEY not found. Add it to .env before running."
#     )

# if not os.getenv("DEEPSEEK_API_KEY"):
#     raise RuntimeError(
#         "DEEPSEEK_API_KEY not found. Add it to .env before running."
#     )

print(
    "Environment variables loaded successfully.",
    # os.getenv("OPENAI_API_KEY"),
    os.getenv("DEEPSEEK_API_KEY"),
)

BASE_DIR = Path(__file__).resolve().parent
PROBLEMS_DIR = str(BASE_DIR / "problems")


RESULTS_FILE = "interview_results.xlsx"

HEADERS = [
    "Candidate Name",
    "Topic",
    "Score",
    "Strengths",
    "Weaknesses",
    "Areas of Improvement",
    "Hired",
]


def save_to_excel(
    candidate_name: str,
    score: int,
    topic: str,
    strengths: str,
    weaknesses: str,
    improvement_areas: str,
    hired: bool,
) -> str:

    if os.path.exists(RESULTS_FILE):
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)

    ws.append([
        candidate_name,
        topic,
        score,
        strengths,
        weaknesses,
        improvement_areas,
        "Yes" if hired else "No",
    ])

    wb.save(RESULTS_FILE)

    return (
        f"Saved interview result for "
        f"{candidate_name} to {RESULTS_FILE}"
    )


async def take_interview(difficulty: str, topic: str):

    if difficulty not in ["Easy", "Medium", "Hard"]:
        print(
            "Invalid difficulty level. "
            "Please choose Easy, Medium, or Hard."
        )
        return

    # client = OpenAIChatCompletionClient(model="gpt-4o")

    # DeepSeek configuration example
    client = OpenAIChatCompletionClient(
        model="deepseek/deepseek-chat",
        base_url="https://openrouter.ai/api/v1",
        api_key=os.getenv("DEEPSEEK_API_KEY"),
        model_info={
            "vision": False,
            "function_calling": True,
            "json_output": True,
            "structured_output": True,
            "family": "unknown",
        },
    )
#     client = OpenAIChatCompletionClient(
#     model="llama-3.3-70b-versatile",
#     base_url="https://api.groq.com/openai/v1",
#     api_key=os.getenv("GROQ_API_KEY"),
#     model_info={
#         "vision": False,
#         "function_calling": True,
#         "json_output": True,
#         "structured_output": True,
#         "family": "llama",
#     },
# )

    fs_params = StdioServerParams(
        command="npx",
        args=[
            "-y",
            "@modelcontextprotocol/server-filesystem",
            PROBLEMS_DIR,
        ],
        read_timeout_seconds=180,
    )

    async with McpWorkbench(fs_params) as fs_tool:

        interviewer_system_message = (
            "You are a technical interviewer for hiring a software engineer.\n\n"

            "INTERVIEW FLOW:\n"
            "1. Greet the candidate and ask for their name.\n"
            "2. Generate EXACTLY ONE coding problem.\n"
            "3. Write the problem into a Python file using the filesystem tool.\n"
            "4. Ask the candidate to solve the problem in that file.\n"
            "5. Review the submitted solution and ask follow-up questions.\n"
            "6. Ask at most 3 follow-up questions total.\n"
            "7. When finished, say exactly 'INTERVIEW_COMPLETE' on its own line.\n\n"

            f"DIFFICULTY:\n"
            f"- The problem difficulty MUST be '{difficulty}'.\n\n"

            f"TOPIC RESTRICTIONS:\n"
            f"- Allowed topics: {topic}\n"
            f"- If 'Any' is NOT present, the generated problem MUST strictly "
            f"belong to ONE of the allowed topics only.\n"
            f"- Do NOT generate problems from unrelated topics.\n"
            f"- Do NOT mix multiple unrelated topics unless naturally required.\n"
            f"- If 'Any' is present, you may choose any standard DSA topic.\n\n"

            "TOPIC EXAMPLES:\n"
            "- Arrays -> sliding window, prefix sums, two pointers, array manipulation\n"
            "- Linked Lists -> reversal, cycle detection, merging, traversal\n"
            "- Stacks and Queues -> monotonic stack, queue simulation, parsing\n"
            "- Trees and Graphs -> DFS, BFS, shortest path, tree traversal\n"
            "- Hashing -> frequency maps, sets, hashmap optimization\n"
            "- Sorting and Searching -> binary search, sorting-based optimization\n"
            "- Dynamic Programming -> memoization, tabulation, state transition\n"
            "- Recursion and Backtracking -> permutations, combinations, recursion\n"
            "- Greedy Algorithms -> interval scheduling, greedy selection\n"
            "- Bit Manipulation -> XOR, masking, bit operations\n\n"

            f"FILESYSTEM RULES:\n"
            f"- Use EXACTLY this directory path:\n"
            f"{PROBLEMS_DIR}\n"
            f"- Never invent, modify, or guess another path.\n"
            f"- Always use the FULL ABSOLUTE PATH when calling the filesystem tool.\n"
            f"- Never use relative paths.\n\n"

            "PROBLEM FILE RULES:\n"
            "- Derive a short snake_case slug from the problem title.\n"
            "- Example slugs:\n"
            "  - longest_substring\n"
            "  - two_sum\n"
            "  - binary_tree_paths\n\n"

            f"- The final file path MUST follow this format:\n"
            f"  {PROBLEMS_DIR}/<slug>.py\n\n"

            "FILE CONTENT RULES:\n"
            "- The ENTIRE problem statement MUST be written as Python comments.\n"
            "- Every problem statement line MUST start with '# '.\n"
            "- Include:\n"
            "  - problem description\n"
            "  - constraints\n"
            "  - input/output explanation\n"
            "  - 1-2 examples\n\n"

            "STARTER CODE RULES:\n"
            "- After the comment block, ALWAYS generate starter code.\n"
            "- ALWAYS include:\n"
            "  - function definition\n"
            "  - correct input parameters\n"
            "  - appropriate return type\n"
            "  - optional type hints\n"
            "  - short docstring\n"
            "  - empty implementation using 'pass'\n\n"

            "STARTER CODE EXAMPLE:\n\n"

            "def two_sum(nums: list[int], target: int) -> list[int]:\n"
            "    \"\"\"Return indices of two numbers whose sum equals target.\"\"\"\n"
            "    pass\n\n"

            "CANDIDATE INSTRUCTIONS:\n"
            "- Announce the exact filename you created.\n"
            "- Tell the candidate to open the generated file.\n"
            "- Tell them to write their solution BELOW the starter function.\n"
            "- Tell them to save the file and reply with 'done'.\n\n"

            "CODE REVIEW RULES:\n"
            "- If the candidate solution has bugs, missing edge cases, "
            "incorrect complexity, or poor design, the Interviewer MUST "
            "ask the candidate to MODIFY the code.\n"
            
            "FOLLOW-UP RULES:\n"
            "- When the candidate replies 'done', read the SAME file back "
            "using the filesystem tool.\n"
            "- Ask focused follow-up questions about:\n"
            "  - edge cases\n"
            "  - time complexity\n"
            "  - space complexity\n"
            "  - optimization\n"
            "  - design decisions\n\n"

            "IMPORTANT:\n"
            "- Never solve the problem for the candidate.\n"
            "- Never give away the final algorithm unless explicitly asked.\n"
            "- Keep follow-up questions concise and technical.\n"
        )
        evaluator_system_message = (
            f"You are a code reviewer. "
            f"Look at the conversation history to find the "
            f"filename the Interviewer announced "
            f"(a '<slug>.py' file inside '{PROBLEMS_DIR}'). "
            f"Use the filesystem tool to read that file using "
            f"its full absolute path "
            f"'{PROBLEMS_DIR}/<slug>.py' — never a relative path. "
            f"The problem statement is in the '#' comment block "
            f"at the top; the candidate's solution is the code "
            f"below it. Produce a structured review:\n"
            "- Correctness: does the solution solve the "
            "stated problem?\n"
            "- Complexity: time and space complexity.\n"
            "- Edge cases: what is handled, what is missed.\n"
            "- Style: readability, naming, structure.\n"
            "Do not address the candidate. "
            "Your output is for the Scorer."
        )

        scorer_system_message = (
            "You are the final scorer. "
            "Using the full conversation and the "
            "CodeEvaluator's review, decide:\n"
            "- topic: the topic of the problem "
            "(e.g. arrays, dynamic programming, etc.)\n"
            "- score: integer out of 10\n"
            "- strengths, weaknesses, improvement_areas: "
            "short strings\n"
            "- hired: true or false\n"
            "Call the save_to_excel tool with the candidate's "
            "name and these fields. "
            "After the tool confirms success, say exactly "
            "'INTERVIEW_FINALIZED' on its own line."
        )

        interviewer_agent = AssistantAgent(
            model_client=client,
            name="Interviewer",
            system_message=interviewer_system_message,
            workbench=fs_tool,
        )

        evaluator_agent = AssistantAgent(
            model_client=client,
            name="CodeEvaluator",
            system_message=evaluator_system_message,
            workbench=fs_tool,
        )

        scorer_agent = AssistantAgent(
            model_client=client,
            name="Scorer",
            system_message=scorer_system_message,
            tools=[save_to_excel],
        )

        candidate = UserProxyAgent(
            name="Candidate",
            input_func=input,
        )

        termination = (
            TextMentionTermination("INTERVIEW_FINALIZED")
            | MaxMessageTermination(60)
        )

        selector_prompt = """
            Select the next speaker carefully.

            SYSTEM GOAL:
            The interview MUST feel like a real technical interview.
            Do NOT rush to finish the interview.
            Do NOT prematurely hand control to the evaluator or scorer.

            ROLES:
            - Interviewer:
                - Creates the coding problem
                - Asks technical follow-up questions
                - May request code changes or optimizations
                - Reviews candidate reasoning
                - Decides when the interview is truly complete
                - ONLY the Interviewer may say 'INTERVIEW_COMPLETE'

            - Candidate:
                - Human participant
                - Answers questions
                - Updates code when requested
                - Explains reasoning and tradeoffs

            - CodeEvaluator:
                - ONLY speaks AFTER Interviewer says 'INTERVIEW_COMPLETE'
                - Reads and evaluates the FINAL code submission
                - Reviews correctness, complexity, edge cases, and style

            - Scorer:
                - ONLY speaks AFTER CodeEvaluator finishes evaluation
                - Calls save_to_excel
                - Says 'INTERVIEW_FINALIZED'

            CRITICAL RULES:
            1. Never select Scorer before CodeEvaluator has completed a review.

            2. Never select CodeEvaluator before Interviewer explicitly says:
            'INTERVIEW_COMPLETE'

            3. If the Interviewer requested:
            - a code change
            - bug fix
            - optimization
            - edge case handling
            - refactor
            - complexity improvement
            then the Candidate MUST respond FIRST.

            4. After the Candidate responds with updated code or explanation,
            control MUST return to the Interviewer.

            5. The Interviewer may continue asking additional follow-ups after
            a code update.

            6. The interview is NOT complete simply because the candidate answered once.

            7. Do NOT rush to evaluation after a partial answer.

            8. If the candidate says:
            - "done"
            - "updated"
            - "fixed"
            - "refactored"
            - "optimized"
            then the next speaker should usually be the Interviewer.

            9. Never select Candidate twice in a row.

            10. The Scorer MUST ONLY appear after:
                Interviewer -> CodeEvaluator -> Scorer

            EXPECTED FLOW:
            Interviewer -> Candidate
            Candidate -> Interviewer
            Interviewer -> Candidate
            Candidate -> Interviewer
            ...
            Interviewer says INTERVIEW_COMPLETE
            -> CodeEvaluator
            -> Scorer

            {roles}

            CONVERSATION HISTORY:
            {history}

            Based on the rules above, select the SINGLE best next speaker from:
            {participants}
            """
        team = SelectorGroupChat(
            participants=[
                interviewer_agent,
                candidate,
                evaluator_agent,
                scorer_agent,
            ],
            model_client=client,
            termination_condition=termination,
            selector_prompt=selector_prompt,
        )

        await Console(
            team.run_stream(
                task=(
                    f"Start a coding interview for a software engineer role. "
                    f"Design the coding problem around the topic(s): {topic}. "
                    f"If the topic contains 'Any', you may choose any DSA topic. "
                    f"Conduct the interview professionally, evaluate the solution, "
                    f"and record the final hiring decision."
                )
            )
        )